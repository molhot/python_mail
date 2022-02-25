import sys
import openpyxl
import smtplib
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

mode = input("モード選択(テスト=test,本番=real):")

if mode != "real":
    test_mode = True
else:
    test_mode = False

if test_mode:
    result = input("テストモードで自分宛てに送信します(続行=yes,中止=no):")
else:
    result = input("本番モードで自分宛てに送信します(続行=yes,中止=no):")

if result != "yes":
    print("プログラムを中止します")
    sys.exit()

else:
    my_address = "sakata19991214@gmail.com"
    smtp_server = "smtp.gmail.com"
    port_number = 587

    account = "sakata19991214@gmail.com"
    password = "jrldrmxavcnmzfwj"

    server = smtplib.SMTP(smtp_server, port_number)
    server.starttls()
    server.login(account, password)

    wb = openpyxl.load_workbook("Book1.xlsx", data_only=True)
    ws = wb["Sheet1"]

    #以降メールアドレスと企業名を順に格納する

    i = 2

    while(ws["A" + str(i)].value != None):
        text = open('企業メール.txt', 'r', encoding = 'utf-8')
        body_temp = text.read()

        customer_name = ws["A" + str(i)].value
        mailaddress = ws["B" + str(i)].value
        customer_tantou = ws["D" + str(i)].value

        msg = MIMEMultipart()
        msg["Subject"] = "ご挨拶"
        msg["From"] = my_address
        msg["To"] = mailaddress

        body_text = body_temp.format(
            company = customer_name,
            person = customer_tantou
        )

        text.close()

        body = MIMEText(body_text)
        msg.attach(body)

        print("メール送信完了" + customer_name)
        server.send_message(msg)

        i = i + 1
server.quit()